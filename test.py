from jieba import cut
from openpyxl import load_workbook


def dictionary_xls_to_txt(filename: str):
    wb = load_workbook(filename)
    ws = wb.active

    # 確定之後來改這個
    ge = ws.iter_cols(min_col=2, max_col=2)
    all_cell = list(ge)[0]

    for cell in all_cell:
        print(cell.value)


dictionary_xls_to_txt('dictionary/3-12外國國名中譯對照表-初稿.xlsx')
